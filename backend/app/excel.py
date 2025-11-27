# app/excel.py
import os
from openpyxl import Workbook, load_workbook
from .config import EXCEL_BASE_DIR

EXCEL_PATH = os.path.join(EXCEL_BASE_DIR, "appointments.xlsx")

def append_appointment_to_excel(appointment):
    """
    appointment: Appointment SQLAlchemy object
    Writes/updates a single Excel file for all appointments.
    """
    # If file exists, load it; otherwise create a new workbook and headers
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"
        ws.append([
            "ID",
            "Business ID",
            "Customer Name",
            "Service",
            "DateTime",
            "Phone",
            "Notes",
            "Created At"
        ])

    ws.append([
        appointment.id,
        appointment.business_id,
        appointment.customer_name,
        appointment.service_type,
        appointment.date_time_str,
        appointment.phone_number,
        appointment.notes or "",
        appointment.created_at.isoformat()
    ])

    wb.save(EXCEL_PATH)
